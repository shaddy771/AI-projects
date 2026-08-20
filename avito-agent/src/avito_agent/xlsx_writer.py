"""Запись XLSX в формате, близком к автозагрузке Авито."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .categories import COMMON_COLUMNS
from .models import ListingRow


INSTRUCTION_LINES = [
    "Файл подготовлен агентом avito-agent для автозагрузки Авито.",
    "Перед загрузкой сверьте колонки с актуальным шаблоном вашей категории:",
    "https://autoload.avito.ru/format/  →  Правила и шаблоны",
    "",
    "Важно:",
    "1. Id должен быть стабильным между выгрузками (не меняйте без нужды).",
    "2. AvitoId оставляйте пустым для новых объявлений; не копируйте чужой AvitoId.",
    "3. ImageUrls — прямые публичные ссылки на фото, разделитель |",
    "4. Либо используйте ImageNames + zip-архив фото (общий размер с xlsx ≤ 100 МБ).",
    "5. Price — целое число в рублях без пробелов и символа валюты.",
    "6. Address — полный адрес так, как он заведён на Авито.",
    "7. После загрузки смотрите отчёт автозагрузки и карточку модерации.",
]


def write_xlsx(
    rows: Iterable[ListingRow],
    output_path: Path,
    *,
    sheet_name: str = "Объявления",
    extra_columns: list[str] | None = None,
) -> Path:
    listings = list(rows)
    columns = list(COMMON_COLUMNS)
    for col in extra_columns or []:
        if col not in columns:
            columns.append(col)
    # подхватить extras из данных
    for row in listings:
        for key in row.extras:
            if key not in columns:
                columns.append(key)

    wb = Workbook()
    instr = wb.active
    instr.title = "Инструкция"
    instr["A1"] = "Инструкция"
    instr["A1"].font = Font(bold=True, size=14)
    for i, line in enumerate(INSTRUCTION_LINES, start=3):
        instr[f"A{i}"] = line
    instr.column_dimensions["A"].width = 100

    ws = wb.create_sheet(sheet_name)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, listing in enumerate(listings, start=2):
        data = listing.to_sheet_dict()
        for col_idx, name in enumerate(columns, start=1):
            value = data.get(name, "")
            if name == "Price" and value != "" and value is not None:
                try:
                    value = int(round(float(value)))
                except (TypeError, ValueError):
                    pass
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if name == "Description":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, name in enumerate(columns, start=1):
        width = 18
        if name in {"Title", "Address"}:
            width = 32
        elif name == "Description":
            width = 55
        elif name in {"ImageUrls", "ImageNames"}:
            width = 45
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(listings) + 1)}"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def rows_to_dicts(rows: Iterable[ListingRow]) -> list[dict[str, Any]]:
    return [r.to_sheet_dict() for r in rows]
